"""
11. Popular Python Packages, 11. Command Query Separation

Command query separation:
Methods or functions should be commands or queries.
Not both.
Commands should make changes, queries should not.
"""
import openpyxl

numbers = [1, 2, 3]

wb = openpyxl.load_workbook("transactions.xlsx")
wb.create_sheet()
sheet = wb["Sheet1"]
sheet.cell()
for row in range(1, 10):
    cell = sheet.cell(row, 1)
    print(cell.value)
sheet.append([1, 2, 3])
wb.save("transactions2.xlsx")
