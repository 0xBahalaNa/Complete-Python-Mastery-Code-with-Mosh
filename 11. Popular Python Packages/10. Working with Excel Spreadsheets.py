"""
11. Popular Python Packages, 10. Working with Excel Spreadsheets

Working with Excel in Python.
Install openpyxl in terminal.
Import openpyxl module.
"""
import openpyxl

# wb = openpyxl.Workbook()
wb = openpyxl.load_workbook("transactions.xlsx")

sheet = wb["Sheet1"]

cell = sheet["a1"]
print(cell.row)
print(cell.column)
print(cell.coordinate)

sheet.cell(row=1, column=1)
for row in range(1, sheet.max_row + 1):
    for column in range(1, sheet.max_column + 1):
        cell = sheet.cell(row, column)
        print(cell.value)

column = sheet["a"]
cells = sheet["a:c"]
sheet["a1:c3"]
sheet[1:3]

sheet.append([1, 2, 3])
sheet.delete_cols

wb.save("transactions2.xlsx")
# wb.create_sheet("Sheet2", 0)
# wb.remove_sheet(sheet)
